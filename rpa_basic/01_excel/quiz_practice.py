from openpyxl import Workbook
wb = Workbook() # 새 워크북 생성
ws = wb.active # 현재 활성화된 sheet 가져옴

ws.append(["학번", "출석", "퀴즈1", "퀴즈2", "중간고사", "기말고사", "프로젝트"])
ws.append([1, 10, 8, 5, 14, 26, 12])
ws.append([2, 7, 3, 7, 15, 24, 18])
ws.append([3, 9, 5, 8, 8, 12, 4])
ws.append([4, 7, 8, 7, 17, 21, 18])
ws.append([5, 7, 8, 7, 16, 25, 15])
ws.append([6, 3, 5, 8, 8, 17, 0])
ws.append([7, 4, 9, 10, 16, 27, 18])
ws.append([8, 6, 6, 6, 15, 19, 17])
ws.append([9, 10, 10, 9, 19, 30, 19])
ws.append([10, 9, 8, 8, 20, 25, 20])

for x in range(2, ws.max_row + 1):
    ws.cell(column=4, row=x, value=10)

ws["H1"] = "총점"
for x in range(2, ws.max_row + 1):
    ws.cell(column=8, row=x, value=f"=SUM(B{x}:G{x})")

ws["I1"] = "성적"
for x in range(2, ws.max_row + 1):
    grade = ws.cell(column=9, row=x)
    if ws.cell(column=2, row=x).value < 5:
        grade.value = "F"
    else:
        score = 0
        for y in range(2, 8):
            score += ws.cell(column=y, row=x).value
            if score >= 90:
                grade.value = "A"
            elif score >= 80:
                grade.value = "B"
            elif score >= 70:
                grade.value = "C"
            else:
                grade.value = "D"

wb.save("rpa_basic\\01_excel\\quiz_practice.xlsx")