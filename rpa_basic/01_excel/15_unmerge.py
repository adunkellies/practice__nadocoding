from openpyxl import load_workbook
wb = load_workbook("rpa_basic\\01_excel\\sample_merge.xlsx")
ws = wb.active

# B2:D2 병합되어 있던 셀을 해제
ws.unmerge_cells("B2:D2")
wb.save("rpa_basic\\01_excel\\sample_unmerge.xlsx")